import logging
from contextlib import contextmanager

import openpyxl
import pandas as pd

from japan.runner import run_validation
from japan.services.checkpoint_service import initialize_clean_checkpoint


class FakePage:
    def __init__(self):
        self.reload_count = 0

    def reload(self):
        self.reload_count += 1


class FakeDrugPage:
    dataset = {}
    no_result_codes = set()

    def __init__(self, page):
        self.page = page
        self.current_drug_id = None

    def search_drug(self, drug_id):
        self.current_drug_id = drug_id

    def is_no_result_page(self):
        return self.current_drug_id in self.no_result_codes

    def open_matching_result(self):
        return self.current_drug_id not in self.no_result_codes

    def extract_details(self, yj_code_from_excel):
        return self.dataset.get(yj_code_from_excel, {})


@contextmanager
def fake_browser_factory():
    yield FakePage()


def make_logger():
    logger = logging.getLogger("runner-test")
    logger.handlers.clear()
    logger.propagate = False
    logger.addHandler(logging.NullHandler())
    return logger


def create_input_workbook(path):
    sheet_one = pd.DataFrame(
        [
            {
                "薬価基準収載医薬品コード": "1111",
                "成分名": "Ingredient One",
                "品名": "Brand One",
                "メーカー名": "Company One",
                "薬価": "10",
            },
            {
                "薬価基準収載医薬品コード": "2222",
                "成分名": "Ingredient Two",
                "品名": "Brand Two",
                "メーカー名": "Company Two",
                "薬価": "20",
            },
        ]
    )
    sheet_two = pd.DataFrame(
        [
            {
                "薬価基準収載医薬品コード": "3333",
                "成分名": "Ingredient Three",
                "品名": "Brand Three",
                "メーカー名": "Company Three",
                "薬価": "30",
            }
        ]
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        sheet_one.to_excel(writer, sheet_name="SheetA", index=False)
        sheet_two.to_excel(writer, sheet_name="SheetB", index=False)


def test_clean_run_creates_csv_checkpoints_and_final_excel(tmp_path, monkeypatch):
    input_file = tmp_path / "input.xlsx"
    checkpoint_root = tmp_path / "checkpoints"
    output_dir = tmp_path / "output"
    create_input_workbook(input_file)

    FakeDrugPage.dataset = {
        "1111": {
            "brand": "Brand One",
            "brand_en": "Brand One EN",
            "company": "Company One",
            "price": "10",
            "ingredient_english": "Ingredient One EN",
            "atc_code": "ATC1111",
        },
        "3333": {
            "brand": "Brand Three",
            "brand_en": "Brand Three EN",
            "company": "Other Company",
            "price": "35",
            "ingredient_english": "",
            "atc_code": "",
        },
    }
    FakeDrugPage.no_result_codes = {"2222"}

    monkeypatch.setattr("japan.runner.translate_company", lambda text: f"{text} EN")

    summary = run_validation(
        mode="clean-run",
        input_file=input_file,
        checkpoint_root=checkpoint_root,
        output_dir=output_dir,
        autosave_every=1,
        browser_factory=fake_browser_factory,
        drug_page_cls=FakeDrugPage,
        logger=make_logger(),
    )

    assert summary["processed"] == 3
    assert summary["all_match"] == 1
    assert summary["mismatched"] == 1
    assert summary["not_found"] == 1
    assert summary["errors"] == 0

    checkpoint_dir = checkpoint_root / summary["checkpoint_name"]
    assert checkpoint_dir.exists()
    assert (checkpoint_dir / "checkpoint_meta.json").exists()
    assert list(checkpoint_dir.glob("sheet_*.csv"))

    workbook = openpyxl.load_workbook(summary["output_file"])
    assert workbook.sheetnames == ["SheetA", "SheetB"]

    sheet_a = workbook["SheetA"]
    headers = [cell.value for cell in sheet_a[1]]
    assert "validation_remarks" in headers
    assert "web_status" in headers

    remarks_col = headers.index("validation_remarks") + 1
    status_col = headers.index("web_status") + 1
    manufacture_col = headers.index("manufacture_name") + 1

    assert sheet_a.cell(2, remarks_col).value == "All Match"
    assert sheet_a.cell(2, status_col).value == "Found"
    assert sheet_a.cell(2, manufacture_col).value == "Company One EN"
    assert sheet_a.cell(3, status_col).value == "Not Found"

    green_rgb = "00C6EFCE"
    red_rgb = "00FFC7CE"
    assert sheet_a.cell(2, remarks_col).fill.start_color.rgb == green_rgb
    assert sheet_a.cell(3, status_col).fill.start_color.rgb == red_rgb


def test_resume_from_named_checkpoint_processes_only_pending_rows(tmp_path, monkeypatch):
    input_file = tmp_path / "input.xlsx"
    checkpoint_root = tmp_path / "checkpoints"
    output_dir = tmp_path / "output"
    create_input_workbook(input_file)

    checkpoint_dir, _, sheets = initialize_clean_checkpoint(
        input_file=input_file,
        base_dir=checkpoint_root,
        checkpoint_name="custom-checkpoint",
    )
    sheets["SheetA"].at[0, "web_status"] = "Found"
    sheets["SheetA"].at[0, "validation_remarks"] = "All Match"
    sheets["SheetA"].at[1, "web_status"] = ""
    sheets["SheetB"].at[0, "web_status"] = ""

    for index, sheet_name in enumerate(sheets.keys(), start=1):
        sheets[sheet_name].to_csv(checkpoint_dir / f"sheet_{index:03d}.csv", index=False, encoding="utf-8-sig")

    FakeDrugPage.dataset = {
        "2222": {
            "brand": "Brand Two",
            "brand_en": "Brand Two EN",
            "company": "Company Two",
            "price": "20",
            "ingredient_english": "Ingredient Two EN",
            "atc_code": "ATC2222",
        },
        "3333": {
            "brand": "Brand Three",
            "brand_en": "Brand Three EN",
            "company": "Company Three",
            "price": "30",
            "ingredient_english": "Ingredient Three EN",
            "atc_code": "ATC3333",
        },
    }
    FakeDrugPage.no_result_codes = set()
    monkeypatch.setattr("japan.runner.translate_company", lambda text: f"{text} EN")

    summary = run_validation(
        mode="resume-from",
        checkpoint_ref="custom-checkpoint",
        checkpoint_root=checkpoint_root,
        output_dir=output_dir,
        autosave_every=10,
        browser_factory=fake_browser_factory,
        drug_page_cls=FakeDrugPage,
        logger=make_logger(),
    )

    assert summary["processed"] == 2
    workbook = openpyxl.load_workbook(summary["output_file"])
    sheet_a = workbook["SheetA"]
    headers = [cell.value for cell in sheet_a[1]]
    status_col = headers.index("web_status") + 1
    remarks_col = headers.index("validation_remarks") + 1

    assert sheet_a.cell(2, status_col).value == "Found"
    assert sheet_a.cell(2, remarks_col).value == "All Match"
    assert sheet_a.cell(3, status_col).value == "Found"


def test_resolve_errors_retries_only_error_rows(tmp_path, monkeypatch):
    input_file = tmp_path / "input.xlsx"
    checkpoint_root = tmp_path / "checkpoints"
    output_dir = tmp_path / "output"
    create_input_workbook(input_file)

    checkpoint_dir, _, sheets = initialize_clean_checkpoint(
        input_file=input_file,
        base_dir=checkpoint_root,
        checkpoint_name="error-checkpoint",
    )
    sheets["SheetA"].at[0, "web_status"] = "Error"
    sheets["SheetA"].at[0, "validation_remarks"] = "Temporary failure"
    sheets["SheetA"].at[1, "web_status"] = "Not Found"
    sheets["SheetB"].at[0, "web_status"] = "Found"
    sheets["SheetB"].at[0, "validation_remarks"] = "All Match"

    for index, sheet_name in enumerate(sheets.keys(), start=1):
        sheets[sheet_name].to_csv(checkpoint_dir / f"sheet_{index:03d}.csv", index=False, encoding="utf-8-sig")

    FakeDrugPage.dataset = {
        "1111": {
            "brand": "Brand One",
            "brand_en": "Brand One EN",
            "company": "Company One",
            "price": "10",
            "ingredient_english": "Ingredient One EN",
            "atc_code": "ATC1111",
        }
    }
    FakeDrugPage.no_result_codes = set()
    monkeypatch.setattr("japan.runner.translate_company", lambda text: f"{text} EN")

    summary = run_validation(
        mode="resolve-errors",
        checkpoint_root=checkpoint_root,
        output_dir=output_dir,
        autosave_every=10,
        browser_factory=fake_browser_factory,
        drug_page_cls=FakeDrugPage,
        logger=make_logger(),
    )

    assert summary["processed"] == 1
    workbook = openpyxl.load_workbook(summary["output_file"])
    sheet_a = workbook["SheetA"]
    headers = [cell.value for cell in sheet_a[1]]
    status_col = headers.index("web_status") + 1
    remarks_col = headers.index("validation_remarks") + 1

    assert sheet_a.cell(2, status_col).value == "Found"
    assert sheet_a.cell(2, remarks_col).value == "All Match"
    assert sheet_a.cell(3, status_col).value == "Not Found"
