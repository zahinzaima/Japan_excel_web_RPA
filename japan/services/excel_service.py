from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from japan import config


COLUMN_MAPPING = {
    "薬価基準収載医薬品コード": [
        "薬価基準収載医薬品コード",
        "薬価基準収載医薬品コード (YJ Code)",
        "YJ Code",
        "Drug price standard listed drug code",
    ],
    "成分名": [
        "成分名",
        "Ingredient name",
    ],
    "品名": [
        "品名",
        "Product name",
    ],
    "メーカー名": [
        "メーカー名",
        "Manufacture name",
    ],
    "薬価": [
        "薬価",
        "Drug price",
    ],
}

VALIDATION_COLUMNS = [
    "validation_remarks",
    "web_status",
    "ingredient",
    "brand_dosage",
    "manufacture_name",
    "atc_code",
]


def normalize_columns(df):
    new_columns = {}

    for standard_col, variations in COLUMN_MAPPING.items():
        for col in df.columns:
            col_clean = str(col).strip().lower()

            for variation in variations:
                if col_clean == variation.strip().lower():
                    new_columns[col] = standard_col
                    break

    return df.rename(columns=new_columns)


def ensure_validation_columns(df):
    for col in VALIDATION_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    return df


def read_workbook(file_path=None, limit=None):
    workbook_path = Path(file_path or config.INPUT_FILE)

    if not workbook_path.exists():
        raise FileNotFoundError(f"Input file not found: {workbook_path}")

    excel_file = pd.ExcelFile(workbook_path)
    all_sheets = {}

    for sheet in excel_file.sheet_names:
        df = pd.read_excel(workbook_path, sheet_name=sheet, keep_default_na=False)
        df = ensure_validation_columns(normalize_columns(df))

        if df.empty:
            continue

        if limit:
            df = df.head(limit)

        all_sheets[sheet] = df

    if not all_sheets:
        raise ValueError("All sheets are empty in the Excel file.")

    return all_sheets


def read_excel(limit=None):
    return read_workbook(limit=limit)


def read_checkpoint_csv(csv_path):
    df = pd.read_csv(csv_path, keep_default_na=False)
    return ensure_validation_columns(normalize_columns(df))


def save_checkpoint_csv(df, csv_path):
    csv_path = Path(csv_path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")


def export_to_excel(all_sheets_dict, output_file, format_file=False):
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in all_sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    if format_file:
        format_excel_output(output_path)

    return output_path


def format_excel_output(output_file):
    output_path = Path(output_file)
    wb = load_workbook(output_path)

    green_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid",
    )
    red_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid",
    )

    for ws in wb.worksheets:
        headers = [cell.value for cell in ws[1]]

        if "validation_remarks" not in headers or "web_status" not in headers:
            continue

        remarks_col = headers.index("validation_remarks") + 1
        status_col = headers.index("web_status") + 1

        for row in range(2, ws.max_row + 1):
            remarks_cell = ws.cell(row=row, column=remarks_col)
            status_cell = ws.cell(row=row, column=status_col)

            if remarks_cell.value == "All Match":
                remarks_cell.fill = green_fill
            elif remarks_cell.value:
                remarks_cell.fill = red_fill

            if status_cell.value == "Found":
                status_cell.fill = green_fill
            elif status_cell.value == "Not Found":
                status_cell.fill = red_fill

    wb.save(output_path)


def save_excel(all_sheets_dict, output_file, format_file=False):
    return export_to_excel(all_sheets_dict, output_file, format_file=format_file)
